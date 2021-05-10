from pandas import read_excel, ExcelWriter, DataFrame, to_datetime
from tkinter import *
from tkinter import ttk
from openpyxl import load_workbook
from tkinter import messagebox
from tkfontchooser import Font
from datetime import datetime
from os import remove, path, mkdir, rename
from locale import currency, setlocale, LC_ALL
from zipfile import ZipFile


class Check:

    def __init__(self):
        self.data_loc = "files\\data\\data.xlsx"
        self.history_sample = DataFrame({"Date": [],
                                         "Amount": [],
                                         "Type": []})

        self.df = DataFrame({"Name": [],
                             "Date": [],
                             "Place": [],
                             "Amount": [],
                             "Phone": []})

    def void_func(self, arg):
        return arg, self

    def check_data(self):
        if not path.exists(self.data_loc):
            if path.exists("files\\data\\auto_backup\\backup.zip"):
                try:
                    with ZipFile("files\\data\\auto_backup\\backup.zip", "r") as backup_zip:
                        backup_zip.extractall()
                except Exception as e:
                    self.void_func(e)
            else:
                self.df.to_excel("files\\data\\data.xlsx")

    def check_history(self):
        df = read_excel(self.data_loc)
        list_names = list(df['Name'])
        if path.exists("files\\data\\history"):
            for i in list_names:
                if not path.exists(f"files\\data\\history\\{i}.xlsx"):
                    self.history_sample.to_excel(f"files\\data\\history\\{i}.xlsx")
        else:
            mkdir("files\\data\\history")

    @staticmethod
    def check_ico():
        if path.exists("files\\ico"):
            if not path.exists("files\\ico\\close.png"):
                with ZipFile("files\\data\\auto_backup\\ico.zip", 'r') as ico_r:
                    ico_r.extract("close.png", path="files\\ico")
            if not path.exists("files\\ico\\ico1.ico"):
                with ZipFile("files\\data\\auto_backup\\ico.zip", 'r') as ico_r:
                    ico_r.extract("ico1.ico", path="files\\ico")
        else:
            mkdir("files\\ico")
            with ZipFile("files\\data\\auto_backup\\ico.zip", 'r') as ico_r:
                ico_r.extractall("files\\ico")

    def check_all(self):
        self.check_data()
        self.check_history()
        self.check_ico()

    def is_data_valid(self):
        d_loc = "files\\data\\data.xlsx"
        try:
            df = read_excel(d_loc)
            self.void_func(df)
            return True
        except Exception as e:
            self.void_func(e)
            return False

    def make_data_valid(self):
        self.df.to_excel(self.data_loc)
        self.check_history()

    def backup_data(self):
        if (path.exists("files\\data\\data.xlsx")) and (self.is_data_valid() is True):
            with ZipFile("files\\data\\auto_backup\\backup.zip", "w") as data_b:
                data_b.write("files\\data\\data.xlsx")


class Main(Tk):

    def __init__(self):
        super().__init__()
        self.FG_DARK = "#E9118F"
        self.FG_LITE = "#CE389C"
        self.BG_WHITE = "ghost white"
        self.BG_DARK = "#1B1B1B"
        self.fileName = "files\\data\\data.xlsx"
        self.IconName = "files\\ico\\ico1.ico"
        self.title("Khata Manager")
        self.bind("<Key>", self.key_binds)
        self.protocol("WM_DELETE_WINDOW", self.close_win)
        self.iconbitmap(bitmap=self.IconName)
        self.geometry("1225x610")
        self.resizable(width=False, height=False)
        self.mainFrame = Frame(self, bg=self.BG_DARK)
        self.mainFrame.pack(fill=BOTH, expand=True)
        self.data = self.get_data()
        self.result_frame = None
        self.search = None
        self.canvas = None
        self.close_image = PhotoImage(file="files\\ico\\close.png")
        # Declaring Title >>
        self.title_frame = Frame(self.mainFrame, bg=self.BG_DARK)
        self.title_label = Label(self.title_frame, text="Khata Manager", bg=self.BG_DARK,
                                 fg=self.BG_WHITE, font=("Roboto", 15, "bold"))
        self.close_btn = Button(self.title_frame, image=self.close_image,
                                bg=self.BG_DARK, border=0, relief=FLAT,
                                activebackground=self.BG_DARK, command=self.close_win)
        self.title_label.bind("<Button-1>", self.show_about_info)
        self.title_label.bind("<Enter>", self.title_enter)
        self.title_label.bind("<Leave>", self.title_exit)

        self.font_set = Font(self.title_label, self.title_label.cget("font"))

        self.info_tab = Frame(self.mainFrame, bg=self.BG_DARK)
        self.Information_label = Label(self.info_tab, text="Entry Info: ", font=("Roboto", 12, "bold"),
                                       fg=self.BG_WHITE, bg=self.BG_DARK, anchor=W)
        self.info_frame = Frame(self.info_tab, bg=self.BG_DARK)
        self.menu = None
        self.search_win = None
        self.search_text = StringVar()
        self.selected_entry = StringVar()

        # Edit window frame >>
        self.edit_frame = Frame(self.mainFrame, width=500, height=650, bg=self.BG_DARK, padx=10)
        self.edit_title_frame = Frame(self.edit_frame, bg=self.BG_DARK, pady=20)
        self.edit_title = Label(self.edit_title_frame, text="Edit: ", font=("Roboto", 12, "bold"),
                                bg=self.BG_DARK, fg=self.BG_WHITE, anchor=W)
        self.edit_entry_frame = Frame(self.edit_frame, bg=self.BG_DARK)
        self.edit_btn_frame = Frame(self.edit_frame, bg=self.BG_DARK)

        # Name >>
        self.edit_name_frame = Frame(self.edit_entry_frame, bg=self.BG_DARK, pady=8)
        self.edit_name_label = Label(self.edit_name_frame, bg=self.BG_DARK, text="Name:    ",
                                     font=("Roboto", 11, "bold"), fg=self.FG_DARK)
        self.edit_name = Entry(self.edit_name_frame, bg=self.FG_LITE, fg=self.BG_WHITE, font=("verdana", 12, 'italic'),
                               border=0)
        # Date >>
        self.edit_date_frame = Frame(self.edit_entry_frame, bg=self.BG_DARK, pady=8)
        self.edit_date_label = Label(self.edit_date_frame, bg=self.BG_DARK, text="Date:    ",
                                     font=("Roboto", 11, "bold"), fg=self.FG_DARK)
        self.edit_date = Entry(self.edit_date_frame, bg=self.FG_LITE, fg=self.BG_WHITE, font=("verdana", 12, 'italic'),
                               border=0)
        # Amount >>
        self.edit_amount_frame = Frame(self.edit_entry_frame, bg=self.BG_DARK, pady=8)
        self.edit_amount_label = Label(self.edit_amount_frame, bg=self.BG_DARK, text="Amount:    ",
                                       font=("Roboto", 11, "bold"), fg=self.FG_DARK)
        self.edit_amount = Entry(self.edit_amount_frame, bg=self.FG_LITE, fg=self.BG_WHITE,
                                 font=("verdana", 12, 'italic'),
                                 border=0)
        # Place >>
        self.edit_place_frame = Frame(self.edit_entry_frame, bg=self.BG_DARK, pady=8)
        self.edit_place_label = Label(self.edit_place_frame, bg=self.BG_DARK, text="Place:    ",
                                      font=("Roboto", 11, "bold"), fg=self.FG_DARK)
        self.edit_place = Entry(self.edit_place_frame, bg=self.FG_LITE, fg=self.BG_WHITE,
                                font=("verdana", 12, 'italic'),
                                border=0)
        # Phone >>
        self.edit_phone_frame = Frame(self.edit_entry_frame, bg=self.BG_DARK, pady=8)
        self.edit_phone_label = Label(self.edit_phone_frame, bg=self.BG_DARK, text="Phone:    ",
                                      font=("Roboto", 11, "bold"), fg=self.FG_DARK)
        self.edit_phone = Entry(self.edit_phone_frame, bg=self.FG_LITE, fg=self.BG_WHITE,
                                font=("verdana", 12, 'italic'),
                                border=0)

        # Making buttons in edit window >>
        self.edit_btn_frame1 = Frame(self.edit_btn_frame, bg=self.BG_DARK, width=406, pady=5)
        self.edit_btn_frame2 = Frame(self.edit_btn_frame, bg=self.BG_DARK, width=406, pady=5)
        self.edit_btn_frame3 = Frame(self.edit_btn_frame, bg=self.BG_DARK, width=406, pady=5)

        self.edit_btn_edit = Button(self.edit_btn_frame1, text="Save changes", font=("Roboto", 12, "bold"),
                                    bg=self.FG_DARK, fg=self.BG_DARK, command=self.edit_btn_pressed,
                                    border=0, padx=5)
        self.edit_btn_history = Button(self.edit_btn_frame1, text="History", font=("Roboto", 12, "bold"),
                                       bg=self.FG_DARK, fg=self.BG_DARK, command=self.show_history,
                                       border=0, padx=5)
        self.edit_btn_delete = Button(self.edit_btn_frame1, text="Delete", font=("Roboto", 12, "bold"),
                                      bg=self.FG_DARK, fg=self.BG_DARK, command=self.delete_entry,
                                      padx=5, border=0)
        self.edit_credit_frame = Frame(self.edit_btn_frame2, bg=self.BG_DARK, pady=6)
        self.edit_debit_frame = Frame(self.edit_btn_frame2, bg=self.BG_DARK, pady=6)

        self.edit_btn_credit = Button(self.edit_credit_frame, text="Credit Amount",
                                      font=("Roboto", 11, "italic"),
                                      bg=self.FG_DARK, fg=self.BG_DARK, command=self.credit_amount,
                                      border=0, padx=5)
        self.edit_btn_debit = Button(self.edit_debit_frame, text="Debit Amount",
                                     font=("Roboto", 11, "italic"),
                                     bg=self.FG_DARK, fg=self.BG_DARK, command=self.debit_amount,
                                     border=0, padx=5)

        self.edit_btn_new_entry = Button(self.edit_btn_frame3, text="Add New Entry",
                                         font=("Roboto", 12, "bold"),
                                         fg=self.BG_WHITE, bg=self.FG_LITE,
                                         command=self.new_entry)

    def key_binds(self, key_all):
        key = key_all.char
        if key == "\x13":
            if self.selected_entry.get() != "":
                self.edit_btn_pressed()
        elif key == "\x06":
            try:
                self.search.focus_set()
            except Exception as e5:
                self.void_func(e5)
        elif key == "\x1b":
            self.close_win()
        elif key == "\x0e":
            self.new_entry()

    def delete_history(self):
        file = self.selected_entry.get()
        file_name = f"files\\data\\history\\{file}.xlsx"
        if path.exists(file_name):
            remove(file_name)

    def show_history(self):

        def canvas_edit(event):
            self.void_func(event)
            canvas.configure(scrollregion=canvas.bbox("all"))

        def on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (int(event.delta) / 120)), "units")

        def close():
            history.unbind_all("<MouseWheel>")
            history.destroy()
            self.canvas.bind_all("<MouseWheel>", self.on_mousewheel)

        history = Toplevel(self)
        history.protocol("WM_DELETE_WINDOW", close)
        history.title("History - {} - Khata Manager".format(self.selected_entry.get()))
        history.config(bg=self.BG_DARK)
        history.geometry("645x510")
        history.grab_set()
        history.wm_attributes("-toolwindow", True)
        history.resizable(width=False, height=False)
        main_frame = Frame(history, bg=self.BG_DARK)
        main_frame.pack(fill=BOTH, expand=True)
        titles_frame = Frame(main_frame, bg=self.BG_DARK, padx=1, width=550)
        titles_frame.pack(side=TOP)
        titles_frame1 = Frame(titles_frame, bg=self.BG_DARK)
        titles_frame1.pack(side=LEFT)
        titles_frame2 = Frame(titles_frame, bg=self.BG_DARK)
        titles_frame2.pack(side=RIGHT)
        Label(titles_frame1, text="Index",
              bg=self.BG_DARK, font=("Roboto", 13, "bold"), pady=2,
              fg=self.BG_WHITE, anchor=CENTER, padx=5+10).pack(side=LEFT)
        Label(titles_frame1, text="Date (dd-mm-yyyy)",
              bg=self.BG_DARK, font=("Roboto", 13, "bold"), pady=2,
              fg=self.BG_WHITE, anchor=CENTER, padx=5+20+8).pack(side=RIGHT)
        Label(titles_frame2, text="Amount",
              bg=self.BG_DARK, font=("Roboto", 13, "bold"), pady=2,
              fg=self.BG_WHITE, anchor=CENTER, padx=5+20+2).pack(side=LEFT)
        Label(titles_frame2, text="Type",
              bg=self.BG_DARK, font=("Roboto", 13, "bold"), pady=2,
              fg=self.BG_WHITE, anchor=CENTER, padx=60).pack(side=RIGHT)
        self.reload()
        history_data = self.get_history()
        data_frame = Frame(main_frame, bg=self.BG_DARK)
        data_frame.pack(fill=BOTH, expand=True)

        style = ttk.Style()
        style.theme_use("clam")
        # configure the style
        style.configure("Vertical.TScrollbar", gripcount=4,
                        background=self.FG_DARK, darkcolor=self.FG_DARK, lightcolor=self.FG_LITE,
                        troughcolor=self.BG_DARK, bordercolor=self.BG_WHITE, arrowcolor=self.BG_DARK)
        scrollbar = ttk.Scrollbar(data_frame, orient=VERTICAL)

        canvas = Canvas(data_frame, bg=self.BG_DARK, border=0, width=625)
        canvas.pack(side=LEFT, fill=BOTH)
        scrollbar.pack(side=RIGHT, fill=Y)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.bind_all("<MouseWheel>", on_mousewheel)
        scrollbar.configure(command=canvas.yview)
        data_main_frame = Frame(canvas,
                                border=0, bg=self.BG_DARK)
        canvas.create_window((0, 0), window=data_main_frame)
        data_main_frame.bind("<Configure>", canvas_edit)

        for index, value in history_data.iterrows():
            frame_data = Frame(data_main_frame, bg=self.BG_DARK)
            frame_data.pack(fill=X)
            frame1 = Frame(frame_data, bg=self.BG_DARK)
            frame1.pack(side=LEFT)
            frame2 = Frame(frame_data, bg=self.BG_DARK)
            frame2.pack(side=RIGHT)
            Label(frame1, text=index,
                  bg=self.BG_DARK, font=("Roboto", 12, "italic"), pady=2,
                  fg=self.FG_DARK, anchor=CENTER, padx=20+15).pack(side=LEFT)
            date = value['Date'].strftime("%d-%m-%Y %H:%M")
            Label(frame1, text=date,
                  bg=self.BG_DARK, font=("Roboto", 12, "italic"), pady=2,
                  fg=self.FG_DARK, anchor=CENTER, padx=10+15).pack(side=RIGHT)
            amount = str(self.add_comma_to_num(float(value['Amount'])))
            Label(frame2, text="Rs " + amount,
                  bg=self.BG_DARK, font=("Roboto", 12, "italic"), pady=2,
                  fg=self.FG_DARK, anchor=CENTER, padx=10+30).pack(side=LEFT)
            Label(frame2, text=value['Type'],
                  bg=self.BG_DARK, font=("Roboto", 12, "italic"), pady=2,
                  fg=self.FG_DARK, anchor=CENTER, padx=10+15).pack(side=RIGHT)

        if len(data_main_frame.winfo_children()) == 0:
            Label(data_main_frame, text="No Credit Or Debit History Found!",
                  bg=self.BG_DARK, font=("Roboto", 14, "italic"), pady=25,
                  fg=self.FG_DARK, anchor=CENTER, padx=10+15).pack(side=TOP)

        history.mainloop()

    @staticmethod
    def add_comma_to_num(num):
        setlocale(LC_ALL, '')
        num = currency(num, symbol=False, grouping=True)
        return num

    @staticmethod
    def show_error(message):
        messagebox.showerror("Error - Khata Manager", message)

    def get_index_of_history(self):
        selected_entry = self.selected_entry.get()
        a = read_excel(f"files\\data\\history\\{selected_entry}.xlsx")
        try:
            a.rename({"Unnamed: 0": "a"}, axis="columns", inplace=True)
            a.drop(["a"], axis=1, inplace=True)
        except Exception as e1:
            self.void_func(e1)
        return int(a.index.size)

    def get_history(self):
        selected_entry = self.selected_entry.get()
        file_name = f"files\\data\\history\\{str(selected_entry)}.xlsx"
        history_df = read_excel(file_name)
        try:
            history_df.rename({"Unnamed: 0": "a"}, axis="columns", inplace=True)
            history_df.drop(["a"], axis=1, inplace=True)
        except Exception as e1:
            self.void_func(e1)
        return history_df

    def get_last_serial_no(self):
        self.reload()
        if len(self.data) != 0:
            last_key = list(self.data.keys())[-1]
            last_serial_no = self.data[last_key]['Serial No.']
        else:
            last_serial_no = 0

        return int(last_serial_no)

    # Removes ews(Extra White Spaces) >>
    @staticmethod
    def remove_ews(string):
        str1 = str(string)
        while True:
            if str1[-1] == " ":
                str1 = str1[0:-1]
            if str1[0] == " ":
                str1 = str1[1:]
            if "  " in str1:
                str1 = str1.replace("  ", " ")
            else:
                break

        return str1

    def close_win(self):
        user_permission = messagebox.askyesno(title="Close Window - Khata Manager",
                                              message="Are you sure to close Khata Manager?")
        if user_permission:
            self.destroy()
            Check("Backup in progress!").backup_data()
            quit(0)
        else:
            pass

    def add_history(self, amount: float, a_type: str):
        date = datetime.now()
        selected_entry = self.selected_entry.get()
        file_name = f"files\\data\\history\\{selected_entry}.xlsx"
        index = self.get_index_of_history()
        df = DataFrame({"": [index],
                        "Date": [date],
                        "Amount": [amount],
                        "Type": [a_type]})
        writer = ExcelWriter(file_name, engine='openpyxl')
        # try to open an existing workbook
        writer.book = load_workbook(file_name)
        # copy existing sheets
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        # read existing file
        reader = read_excel(file_name)
        # write out the new sheet
        df.to_excel(writer, index=False, header=False, startrow=len(reader) + 1)
        writer.close()

    def credit_amount(self):

        def credit_btn_pressed(a=''):
            self.void_func(a)
            amount_to_add = amount_entry.get()
            while True:
                try:
                    amount_to_add = float(self.remove_comma(amount_to_add))
                except ValueError:
                    self.show_error("Amount must be a decimal number or a integer! "
                                    "\nExample: '100', 20.45'...etc.")
                    break

                try:
                    actual_amount = float(self.remove_comma(self.data[self.selected_entry.get()]['Amount']))
                    amount_after_credit = float(amount_to_add + actual_amount)
                    num_index = self.data[self.selected_entry.get()]["Serial No."]
                    self.edit_one_cell(heading_index=5, num_index=int(num_index), value=amount_after_credit)
                    try:
                        self.add_history(float(amount_to_add), a_type="credit")
                    except Exception as e6:
                        self.void_func(e6)
                    credit.destroy()
                    self.search.delete(0, END)
                    self.reload()
                    self.add_all()
                    self.reload_info()
                    break
                except Exception as e:
                    if str(e) == "int too big to convert":
                        self.show_error("Integer is too big!")

                    break

        credit = Toplevel(self)
        credit.bind("<Return>", credit_btn_pressed)
        credit.title("Credit Amount - Khata Manager")
        credit.config(bg=self.BG_DARK)
        credit.geometry("400x187")
        credit.wm_attributes("-toolwindow", True)
        credit.resizable(width=False, height=False)
        credit.grab_set()
        credit_main_frame = Frame(credit, bg=self.BG_DARK)
        credit_main_frame.pack(fill=BOTH, expand=True)
        selected_entry = self.selected_entry.get()
        entry_name_frame = Frame(credit_main_frame, bg=self.BG_DARK)
        entry_name_frame.pack(side=TOP, fill=X)
        entry_name = Label(entry_name_frame, text=f"Credit To: ",
                           bg=self.BG_DARK, fg=self.BG_WHITE,
                           anchor=W, font=("Roboto", 14, "bold"), padx=20)
        entry_name.pack(fill=X, side=LEFT)
        entry_name1 = Label(entry_name_frame, text=selected_entry.capitalize(),
                            bg=self.BG_DARK, fg=self.FG_LITE,
                            anchor=W, font=("Roboto", 13, "bold"))
        entry_name1.pack(fill=X)

        amount_frame = Frame(credit_main_frame, bg=self.BG_DARK)
        amount_frame.pack(fill=X)
        amount_entry_frame = Frame(amount_frame, bg=self.BG_DARK, pady=20)
        amount_entry_frame.pack(side=TOP)
        amount_entry_label = Label(amount_entry_frame, text="Enter Amount To be Credited",
                                   font=("verdana", 12, "italic"), fg=self.FG_DARK, bg=self.BG_DARK,
                                   padx=10, pady=7)
        amount_entry_label.pack(side=TOP)
        amount_entry = Entry(amount_entry_frame, bg=self.FG_LITE, fg=self.BG_WHITE,
                             font=("Roboto", 13, "bold"), border=0)
        amount_entry.pack()
        amount_credit_btn = Button(credit_main_frame, bg=self.FG_DARK, fg=self.BG_WHITE,
                                   text="Credit Amount", font=("Roboto", 12, "bold"), border=0,
                                   command=credit_btn_pressed)
        amount_credit_btn.place(x=60, y=140)
        amount_close_btn = Button(credit_main_frame, bg=self.FG_DARK, fg=self.BG_WHITE,
                                  text="Close", font=("Roboto", 12, "bold"), command=credit.destroy,
                                  border=0)
        amount_close_btn.place(x=260, y=140)

    def debit_amount(self):
        def debit_btn_pressed(a=''):
            self.void_func(a)
            amount_to_add = amount_entry.get()
            while True:
                try:
                    amount_to_subtract = float(self.remove_comma(amount_to_add))
                except ValueError:
                    self.show_error("Amount must be a decimal number or a integer! "
                                    "\nExample: '100', 20.45'...etc.")
                    break

                try:
                    actual_amount = float(self.remove_comma(self.data[self.selected_entry.get()]['Amount']))
                    if actual_amount < amount_to_subtract:
                        self.show_error("Debit Amount Should Not Be Greater Than The Actual Amount!")
                        break
                    amount_after_debit = float(actual_amount - amount_to_subtract)
                    num_index = self.data[self.selected_entry.get()]["Serial No."]
                    self.edit_one_cell(heading_index=5, num_index=int(num_index), value=amount_after_debit)
                    try:
                        self.add_history(float(amount_to_subtract), a_type="debit")
                    except Exception as e1:
                        self.void_func(e1)
                    debit.destroy()
                    self.search.delete(0, END)
                    self.reload()
                    self.add_all()
                    self.reload_info()
                    break
                except Exception as e:
                    print(e)
                    self.void_func(e)
                    break

        debit = Toplevel(self)
        debit.bind("<Return>", debit_btn_pressed)
        debit.title("debit Amount - Khata Manager")
        debit.config(bg=self.BG_DARK)
        debit.geometry("400x187")
        debit.wm_attributes("-toolwindow", True)
        debit.resizable(width=False, height=False)
        debit.grab_set()
        debit_main_frame = Frame(debit, bg=self.BG_DARK)
        debit_main_frame.pack(fill=BOTH, expand=True)
        selected_entry = self.selected_entry.get()
        entry_name_frame = Frame(debit_main_frame, bg=self.BG_DARK)
        entry_name_frame.pack(side=TOP, fill=X)
        entry_name = Label(entry_name_frame, text=f"Debit From: ",
                           bg=self.BG_DARK, fg=self.BG_WHITE,
                           anchor=W, font=("Roboto", 14, "bold"), padx=20)
        entry_name.pack(fill=X, side=LEFT)
        entry_name1 = Label(entry_name_frame, text=selected_entry.capitalize(),
                            bg=self.BG_DARK, fg=self.FG_LITE,
                            anchor=W, font=("Roboto", 13, "bold"))
        entry_name1.pack(fill=X)

        amount_frame = Frame(debit_main_frame, bg=self.BG_DARK)
        amount_frame.pack(fill=X)
        amount_entry_frame = Frame(amount_frame, bg=self.BG_DARK, pady=20)
        amount_entry_frame.pack(side=TOP)
        amount_entry_label = Label(amount_entry_frame, text="Enter Amount To be debited",
                                   font=("verdana", 12, "italic"), fg=self.FG_DARK, bg=self.BG_DARK,
                                   padx=10, pady=7)
        amount_entry_label.pack(side=TOP)
        amount_entry = Entry(amount_entry_frame, bg=self.FG_LITE, fg=self.BG_WHITE,
                             font=("Roboto", 13, "bold"), border=0)
        amount_entry.pack()
        amount_debit_btn = Button(debit_main_frame, bg=self.FG_DARK, fg=self.BG_WHITE,
                                  text="Debit Amount", font=("Roboto", 12, "bold"), border=0,
                                  command=debit_btn_pressed)
        amount_debit_btn.place(x=60, y=140)
        amount_close_btn = Button(debit_main_frame, bg=self.FG_DARK, fg=self.BG_WHITE,
                                  text="Close", font=("Roboto", 12, "bold"), command=debit.destroy,
                                  border=0)
        amount_close_btn.place(x=260, y=140)

    @staticmethod
    def make_history_file(name):
        history_sample = {"Date": [],
                          "Amount": [], "Type": []}
        df = DataFrame(history_sample)
        file_name = f"files\\data\\history\\{name}.xlsx"
        df.to_excel(file_name)

    # Removes aws(All White Spaces) >>
    @staticmethod
    def remove_aws(string):
        string = str(string).replace(" ", "")
        return string

    def new_entry(self):

        def new_add_pressed(a=""):
            self.void_func(a)
            name_ = name.get()
            if name_ == "":
                self.show_error("please fill required entry: Name")
            amount_ = self.remove_ews(self.remove_comma(amount.get()))
            if amount_ == "":
                self.show_error("please fill required entry: Amount")
            place_ = self.remove_ews(place.get())
            if place_ == "":
                self.show_error("please fill required entry: Place")
            date_ = date.get()
            if date_ == "":
                self.shgow_error("please fill required entry: Date")

            try:
                date_ = self.remove_ews(str(date_)).split(" ")
                date_ = str(date_[2]) + "/" + str(date_[1]) + "/" + str(date_[0])
            except IndexError:
                try:
                    date_ = self.remove_ews(str(date_[0])).split("-")
                    date_ = str(date_[2]) + "/" + str(date_[1]) + "/" + str(date_[0])
                except IndexError:
                    if type(date_) == list:
                        date_ = self.remove_ews(date_[0])
                    else:
                        date_ = self.remove_ews(date_)

            phone_ = self.remove_aws(phone.get())

            serial_no_ = self.get_last_serial_no()

            while True:
                try:
                    if (name_.lower() == "nan") or (name_.lower() == "none"):
                        self.show_error("Name Can't be NaN or None!")
                        break
                    if (place_.lower() == "nan") or (place_.lower() == "none"):
                        self.show_error("Place Can't be NaN or None!")
                        break
                    try:
                        phone_ = int(phone_)
                    except Exception as e5:
                        self.void_func(e5)
                        phone_ = "NaN"
                    amount_ = float(amount_)
                    date_ = to_datetime(date_)
                except Exception as phone_e:
                    self.show_error(phone_e)
                    break

                try:
                    if str(name_) in list(self.data.keys()):
                        self.show_error(f"One Entry Already Exists With Name: {str(name_)}")
                        break
                    self.append_data(serial_no=serial_no_, amount=amount_, date=date_,
                                     name=name_, place=place_, phone=phone_)
                    self.make_history_file(name_)
                    self.search.delete(0, END)
                    self.reload()
                    self.add_all()
                    self.reload_info()
                    new_win.destroy()
                    break
                except Exception as append_data_e:
                    self.void_func(append_data_e)
                    new_win.destroy()
                    break

        new_win = Toplevel(self)
        new_win.bind("<Return>", new_add_pressed)
        new_win.title("New Entry - Khata Manager")
        new_win.config(bg=self.BG_DARK)
        new_win.geometry("400x300")
        new_win.wm_attributes("-toolwindow", True)
        new_win.resizable(width=False, height=False)
        new_win.grab_set()
        new_win_main_frame = Frame(new_win, bg=self.BG_DARK)
        new_win_main_frame.pack(fill=BOTH, expand=True)

        new_entry_frame = Frame(new_win_main_frame, bg=self.BG_DARK)

        # Name >>
        name_frame = Frame(new_entry_frame, bg=self.BG_DARK, pady=8)
        name_label = Label(name_frame, bg=self.BG_DARK, text="Name:    ",
                           font=("Roboto", 11, "bold"), fg=self.FG_DARK)
        name = Entry(name_frame, bg=self.FG_LITE, fg=self.BG_WHITE, font=("verdana", 12, 'italic'),
                     border=0)
        # Date >>
        date_frame = Frame(new_entry_frame, bg=self.BG_DARK, pady=8)
        date_label = Label(date_frame, bg=self.BG_DARK, text="Date:    ",
                           font=("Roboto", 11, "bold"), fg=self.FG_DARK)
        date = Entry(date_frame, bg=self.FG_LITE, fg=self.BG_WHITE, font=("verdana", 12, 'italic'),
                     border=0)
        date.insert(0, self.date_now_normal_format())
        # Amount >>
        amount_frame = Frame(new_entry_frame, bg=self.BG_DARK, pady=8)
        amount_label = Label(amount_frame, bg=self.BG_DARK, text="Amount:    ",
                             font=("Roboto", 11, "bold"), fg=self.FG_DARK)
        amount = Entry(amount_frame, bg=self.FG_LITE, fg=self.BG_WHITE,
                       font=("verdana", 12, 'italic'),
                       border=0)
        # Place >>
        place_frame = Frame(new_entry_frame, bg=self.BG_DARK, pady=8)
        place_label = Label(place_frame, bg=self.BG_DARK, text="Place:    ",
                            font=("Roboto", 11, "bold"), fg=self.FG_DARK)
        place = Entry(place_frame, bg=self.FG_LITE, fg=self.BG_WHITE,
                      font=("verdana", 12, 'italic'),
                      border=0)
        # Phone >>
        phone_frame = Frame(new_entry_frame, bg=self.BG_DARK, pady=8)
        phone_label = Label(phone_frame, bg=self.BG_DARK, text="Phone:    ",
                            font=("Roboto", 11, "bold"), fg=self.FG_DARK)
        phone = Entry(phone_frame, bg=self.FG_LITE, fg=self.BG_WHITE,
                      font=("verdana", 12, 'italic'),
                      border=0)

        new_entry_frame.pack(side=TOP)

        # Packing all entries >>
        name_frame.pack(fill=X)
        name_label.pack(side=LEFT)
        name.pack(side=RIGHT)

        date_frame.pack(fill=X)
        date_label.pack(side=LEFT)
        date.pack(side=RIGHT)

        amount_frame.pack(fill=X)
        amount_label.pack(side=LEFT)
        amount.pack(side=RIGHT)

        place_frame.pack(fill=X)
        place_label.pack(side=LEFT)
        place.pack(side=RIGHT)

        phone_frame.pack(fill=X)
        phone_label.pack(side=LEFT)
        phone.pack(side=RIGHT)

        new_btn_frame = Frame(new_win_main_frame, bg=self.BG_DARK)
        new_btn_frame.pack(side=BOTTOM)
        new_add_btn = Button(new_btn_frame, text="Add  ", font=("Roboto", 14, "bold"),
                             bg=self.BG_DARK, fg=self.BG_WHITE, border=0, activebackground=self.BG_WHITE,
                             activeforeground=self.FG_DARK, padx=20, command=new_add_pressed)
        new_close_btn = Button(new_btn_frame, text="Close", font=("Roboto", 14, "bold"),
                               bg=self.BG_DARK, fg=self.BG_WHITE, border=0, activebackground=self.BG_WHITE,
                               activeforeground=self.FG_DARK, padx=20, command=new_win.destroy)

        new_add_btn.pack(side=LEFT)
        new_close_btn.pack(side=RIGHT)

        new_win.mainloop()

    @staticmethod
    def date_now_normal_format():
        date = datetime.now()
        return date.strftime("%d %m %Y")

    def delete_entry(self):
        serial_no = int(self.data[self.selected_entry.get()]["Serial No."])
        df = read_excel(self.fileName)
        try:
            df.rename({"Unnamed: 0": "a"}, axis="columns", inplace=True)
            df.drop(["a"], axis=1, inplace=True)
        except Exception as e1:
            self.void_func(e1)
        user_permission = messagebox.askyesno("Delete Entry - Khata Manager",
                                              f"Are you sure to delete entry? \n"
                                              f"All history and data will be deleted..."
                                              f" \nEntry Name: {self.selected_entry.get()}")
        if user_permission is True:
            df.drop(index=serial_no, inplace=True)
            for i, value in enumerate(df['s. no.']):
                df.iat[i, 0] = i
            df.to_excel(self.fileName)

            df = read_excel(self.fileName)
            try:
                df.rename({"Unnamed: 0": "a"}, axis="columns", inplace=True)
                df.drop(["a"], axis=1, inplace=True)
            except Exception as e1:
                self.void_func(e1)
            df.to_excel(self.fileName)

            for i in self.info_frame.winfo_children():
                i.destroy()
            Label(self.info_frame, text="_-_Choose a Entry_-_", fg=self.FG_DARK,
                  bg=self.BG_DARK, font=("verdana", 14, "bold")).pack(fill=BOTH)

            # Making normal all entries >>
            self.edit_name.delete(0, END)
            self.edit_name['state'] = DISABLED
            self.edit_date.delete(0, END)
            self.edit_date['state'] = DISABLED
            self.edit_place.delete(0, END)
            self.edit_place['state'] = DISABLED
            self.edit_amount.delete(0, END)
            self.edit_amount['state'] = DISABLED
            self.edit_phone.delete(0, END)
            self.edit_phone['state'] = DISABLED
            self.edit_btn_edit['state'] = DISABLED
            self.edit_btn_delete['state'] = DISABLED
            self.edit_btn_credit['state'] = DISABLED
            self.edit_btn_debit['state'] = DISABLED
            self.edit_btn_history['state'] = DISABLED
            self.delete_history()
            self.reload()
            self.add_all()
            self.selected_entry.set("")

        else:
            pass

    def edit_btn_pressed(self):
        while True:
            selected_entry = self.selected_entry.get()
            entry_index = int(self.data[selected_entry]["Serial No."])
            name = self.edit_name.get()
            self.edit_one_cell(heading_index=2, num_index=entry_index, value=str(name))
            try:
                rename(f"files\\data\\history\\{selected_entry}.xlsx", f"files\\data\\history\\{str(name)}.xlsx")
            except Exception as e2:
                self.void_func(e2)
            date = self.edit_date.get()
            try:
                date = self.remove_ews(str(date)).split(" ")
                date = str(date[2]) + "/" + str(date[1]) + "/" + str(date[0])
            except IndexError:
                try:
                    date = self.remove_ews(str(date[0])).split("-")
                    date = str(date[2]) + "/" + str(date[1]) + "/" + str(date[0])
                except IndexError:
                    if type(date) == list:
                        date = self.remove_ews(date[0])
                    else:
                        date = self.remove_ews(date)
            try:
                date = to_datetime(int(date.replace("/", "")), format="%Y%m%d")
            except ValueError:
                self.show_error("Enter Date Correctly! \nFormat out of range!")
                break
            self.edit_one_cell(heading_index=3, num_index=entry_index, value=date)
            amount = self.edit_amount.get()
            amount = self.remove_comma(amount)
            amount_before = float(self.remove_comma(self.data[self.selected_entry.get()]['Amount']))
            if amount > amount_before:
                amount_type = "credit"
                self.add_history(amount=float(amount) - amount_before, a_type=amount_type)
            elif amount < amount_before:
                amount_type = "debit"
                self.add_history(amount=amount_before - float(amount), a_type=amount_type)
            else:
                pass
            self.edit_one_cell(heading_index=5, num_index=entry_index, value=float(amount))
            place = self.edit_place.get()
            self.edit_one_cell(heading_index=4, num_index=entry_index, value=str(place))
            phone = self.remove_aws(self.edit_phone.get())
            if str(phone).lower() != "":
                try:
                    phone = int(phone)
                except Exception as e3:
                    self.void_func(e3)
                    self.show_error("Phone Number must be a integer!")
                    break

            self.edit_one_cell(heading_index=6, num_index=entry_index, value=str(phone))
            self.search.delete(0, END)
            self.reload()
            self.selected_entry.set(str(name))
            self.add_all()
            self.reload_info()
            break

    @staticmethod
    def remove_comma(num_str):
        if "," in str(num_str):
            num_str = str(num_str).replace(",", "")

        return float(num_str)

    def title_enter(self, arg):
        self.void_func(arg)
        self.title_label.config(fg=self.FG_DARK)

    def title_exit(self, arg):
        self.void_func(arg)
        self.title_label.config(fg=self.BG_WHITE)

    def make_title(self):
        self.title_frame.pack(side=TOP, fill=X)
        self.title_label.pack()
        self.close_btn.place(x=1170, y=-5)
        self.font_set.configure(underline=True)
        self.title_label.configure(font=self.font_set)

    def void_func(self, arg):
        return self, arg

    def reload(self):
        self.data = self.get_data()

    def make_edit_frame(self):
        self.info_tab.place(x=520, y=300)
        self.Information_label.pack(side=TOP, fill=X)
        self.info_frame.pack(side=BOTTOM)
        Label(self.info_frame, text="_-_Choose a Entry_-_", fg=self.FG_DARK,
              bg=self.BG_DARK, font=("verdana", 14, "bold")).pack(fill=BOTH)

    def reload_info(self):
        entry_data = self.data[self.edit_name.get()]
        for i1 in self.info_frame.winfo_children():
            i1.destroy()
        if len(self.selected_entry.get()) != 0:
            for i in entry_data.keys():
                element_frame = Frame(self.info_frame, bg=self.BG_DARK)
                element_frame.pack(fill=X)
                if "Amount" in i:
                    i_ = i + " (Rs)"
                    entry_data[i] = self.add_comma_to_num(self.remove_comma(entry_data[i]))
                else:
                    i_ = i
                Label(element_frame, text=f"{str(i_)}:   ", bg=self.BG_DARK,
                      fg=self.FG_DARK, anchor=W, font=("Roboto", 13, 'bold')).pack(side=LEFT)
                if "Date" in i:
                    info_data = str(entry_data[i][0]) + "-" + str(entry_data[i][1]) + "-" + str(entry_data[i][2])
                else:
                    info_data = str(entry_data[i])
                Label(element_frame, text=info_data, bg=self.BG_DARK,
                      fg=self.FG_LITE, anchor=W, font=("Roboto", 12, 'italic')).pack(side=RIGHT)
                if str(i) == "Name":
                    self.edit_name.delete(0, END)
                    self.edit_name.insert(0, entry_data[i])

                if str(i) == "Date":
                    self.edit_date.delete(0, END)
                    self.edit_date.insert(0, entry_data[i])

                if str(i) == "Amount":
                    self.edit_amount.delete(0, END)
                    self.edit_amount.insert(0, entry_data[i])

                if str(i) == "Place":
                    self.edit_place.delete(0, END)
                    self.edit_place.insert(0, entry_data[i])

                if str(i) == "Phone":
                    self.edit_phone.delete(0, END)
                    self.edit_phone.insert(0, entry_data[i])
        else:
            Label(self.info_frame, text="_-_Choose a Entry_-_", fg=self.FG_DARK,
                  bg=self.BG_DARK, font=("verdana", 14, "bold")).pack(fill=BOTH)

    def select(self):
        selected_entry_name = self.selected_entry.get()
        entry_data = self.data[selected_entry_name]
        for i1 in self.info_frame.winfo_children():
            i1.destroy()

        # Making normal all entries >>
        self.edit_name['state'] = NORMAL
        self.edit_date['state'] = NORMAL
        self.edit_amount['state'] = NORMAL
        self.edit_place['state'] = NORMAL
        self.edit_phone['state'] = NORMAL
        self.edit_btn_edit['state'] = NORMAL
        self.edit_btn_delete['state'] = NORMAL
        self.edit_btn_credit['state'] = NORMAL
        self.edit_btn_debit['state'] = NORMAL
        self.edit_btn_history['state'] = NORMAL

        for i in entry_data.keys():
            element_frame = Frame(self.info_frame, bg=self.BG_DARK)
            element_frame.pack(fill=X)
            if "Amount" in i:
                i_ = i + " (Rs)"
                entry_data[i] = self.add_comma_to_num(self.remove_comma(entry_data[i]))
            else:
                i_ = i
            Label(element_frame, text=f"{str(i_)}:   ", bg=self.BG_DARK,
                  fg=self.FG_DARK, anchor=W, font=("Roboto", 13, 'bold')).pack(side=LEFT)
            if "Date" in i:
                info_data = str(entry_data[i][0]) + "-" + str(entry_data[i][1]) + "-" + str(entry_data[i][2])
            else:
                info_data = str(entry_data[i])
            Label(element_frame, text=info_data, bg=self.BG_DARK,
                  fg=self.FG_LITE, anchor=W, font=("Roboto", 12, 'italic')).pack(side=RIGHT)
            if str(i) == "Name":
                self.edit_name.delete(0, END)
                self.edit_name.insert(0, entry_data[i])

            if str(i) == "Date":
                self.edit_date.delete(0, END)
                self.edit_date.insert(0, entry_data[i])

            if str(i) == "Amount":
                self.edit_amount.delete(0, END)
                self.edit_amount.insert(0, entry_data[i])

            if str(i) == "Place":
                self.edit_place.delete(0, END)
                self.edit_place.insert(0, entry_data[i])

            if str(i) == "Phone":
                self.edit_phone.delete(0, END)
                self.edit_phone.insert(0, entry_data[i])

    def add_all(self):
        for i1 in self.result_frame.winfo_children():
            i1.destroy()
        for i in self.data.keys():
            label_text = i + "      " + "Rs " + str(self.add_comma_to_num(float(
                self.remove_comma(self.data[i]["Amount"]))))
            Radiobutton(self.result_frame, value=i, text=label_text,
                        bg=self.BG_DARK, fg=self.FG_DARK, variable=self.selected_entry,
                        font=("Roboto", 12, "bold"), pady=3, anchor=W,
                        border=0, command=self.select,
                        activebackground=self.BG_DARK,
                        activeforeground=self.FG_DARK).pack(fill=X, side=TOP)
        Button(self.result_frame, text="Add New Entry", font=("verdana", 13, "bold"),
               command=self.new_entry, bg=self.BG_DARK, fg=self.BG_WHITE,
               border=0, anchor=W, relief=FLAT, activebackground=self.BG_DARK,
               activeforeground=self.BG_WHITE, pady=14).pack(fill=X, side=BOTTOM)

    def on_mousewheel(self, event):
        self.canvas.yview_scroll(int(-1 * (int(event.delta) / 120)), "units")

    def make_search_window(self):

        def add_result(key):

            for i in self.result_frame.winfo_children():
                i.destroy()

            entry_text = self.search.get()
            for i in self.data.keys():
                if entry_text.lower() in i.lower():
                    label_text = i + "      " + "Rs" + str(self.add_comma_to_num(float(
                        self.remove_comma(self.data[i]["Amount"]))))
                    Radiobutton(self.result_frame, value=i, text=label_text,
                                bg=self.BG_DARK, fg=self.FG_DARK, variable=self.selected_entry,
                                font=("Roboto", 12, "bold"), pady=3, anchor=W,
                                command=self.select,
                                activebackground=self.BG_DARK,
                                activeforeground=self.FG_DARK).pack(fill=X, side=TOP)
            Button(self.result_frame, text="Add New Entry", font=("verdana", 13, "bold"),
                   command=self.new_entry, bg=self.BG_DARK, fg=self.BG_WHITE,
                   border=0, anchor=W, pady=14, activebackground=self.BG_DARK,
                   activeforeground=self.BG_WHITE).pack(fill=X, side=BOTTOM)
            if len(self.result_frame.winfo_children()) == 1:
                Label(self.result_frame, text="No entry found", bg=self.BG_DARK, fg=self.FG_DARK,
                      font=("Roboto", 12, "bold"), pady=3, anchor=W).pack(fill=X, side=TOP)

            return key

        def canvas_edit(event):
            self.void_func(event)
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))

        self.search_win = Frame(self.mainFrame, bg=self.BG_DARK, height=650, width=500)
        self.search_win.place(x=3, y=10)

        # Search Area
        search_frame = Frame(self.search_win, bg=self.BG_DARK, width=360, height=150)
        search_frame.place(x=57, y=50)
        search_label = Label(search_frame, text="Search box", anchor=W,
                             font=("verdana", 11, "italic"), bg=self.BG_DARK, fg=self.FG_DARK)
        search_label.pack(fill=X)
        self.search = Entry(search_frame, bg=self.FG_DARK, fg=self.BG_DARK,
                            font=("Roboto", 13, "italic"), border=0)

        # Result Area
        result_main = Frame(self.search_win, bg=self.BG_DARK, width=400, height=460, border=0)
        result_main.place(x=57, y=170)

        # setting style obj. for the scroll bar
        style = ttk.Style()
        style.theme_use("clam")
        # configure the style
        style.configure("Vertical.TScrollbar", gripcount=4,
                        background=self.FG_DARK, darkcolor=self.FG_DARK, lightcolor=self.FG_LITE,
                        troughcolor=self.BG_DARK, bordercolor=self.BG_WHITE, arrowcolor=self.BG_DARK)
        scrollbar = ttk.Scrollbar(result_main, orient=VERTICAL)

        self.canvas = Canvas(result_main, bg=self.BG_DARK, width=400, height=410, border=0)
        self.canvas.pack(side=LEFT)
        scrollbar.pack(side=RIGHT, fill=Y)
        self.canvas.configure(yscrollcommand=scrollbar.set)
        self.canvas.bind_all("<MouseWheel>", self.on_mousewheel)
        scrollbar.configure(command=self.canvas.yview)
        self.search.bind_all("<Key>", add_result)
        self.search.pack(fill=X)

        self.result_frame = Frame(self.canvas, height=365,
                                  border=0, width=440, bg=self.BG_DARK)
        self.canvas.create_window((0, 0), window=self.result_frame)
        self.result_frame.bind("<Configure>", canvas_edit)
        self.add_all()

    def append_data(self, serial_no, name, date, place, amount, phone):
        df = DataFrame({"": serial_no,
                        "s. no.": [serial_no],
                        "Name": [name],
                        "Date": [date],
                        "Place": [str(place)],
                        "Amount": [float(amount)],
                        "Phone no.": [phone]})
        writer = ExcelWriter(self.fileName, engine='openpyxl')
        # try to open an existing workbook
        writer.book = load_workbook(self.fileName)
        # copy existing sheets
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        # read existing file
        reader = read_excel(self.fileName)
        # write out the new sheet
        for i, value in enumerate(df['s. no.']):
            df.iat[i, 0] = i
        df.to_excel(writer, index=False, header=False, startrow=len(reader) + 1)
        writer.close()
        df = read_excel(self.fileName)
        try:
            df.rename({"Unnamed: 0": "a"}, axis="columns", inplace=True)
            df.drop(["a"], axis=1, inplace=True)
        except Exception as e1:
            self.void_func(e1)
        for i, value in enumerate(df['s. no.']):
            df.iat[i, 0] = i
        df.to_excel(self.fileName)

    def get_data(self):
        raw_data = read_excel(io=self.fileName)
        data = {}
        for index, item in raw_data.iterrows():
            name = item["Name"]
            date = str(item['Date'].strftime(f"%d-%m-%Y")).split("-")
            amount = item["Amount"]
            place = item["Place"]
            if str(item["Phone no."]) != "nan":
                phone = int(item["Phone no."])
            else:
                phone = item["Phone no."]
            serial_no = item["s. no."]
            pre_dict = {"Serial No.": serial_no, "Name": name, "Date": date,
                        "Amount": float(amount), "Place": str(place),
                        "Phone": phone}
            data[name] = pre_dict

        return data

    def edit_one_cell(self, heading_index, num_index: int, value):
        data_frame = read_excel(self.fileName)
        try:
            data_frame.rename({"Unnamed: 0": "a"}, axis="columns", inplace=True)
            data_frame.drop(["a"], axis=1, inplace=True)
        except Exception as e1:
            self.void_func(e1)
        data_frame.iat[int(num_index), int(heading_index)-1] = value
        data_frame.to_excel(self.fileName)

    def show_about_info(self, n):
        self.void_func(n)
        messagebox.showinfo(title="About - Khata Manager",
                            message="Khata Manager is a Account Management System \n"
                                    "\nYou can add accounts of your clients and \nmanage"
                                    " there credits and debits with accurate dates. "
                                    "\nWith fast searching by name you can search your"
                                    " \nclient details faster.")

    def edit_tab(self):
        self.edit_frame.place(y=110, x=838)
        self.edit_title_frame.pack(side=TOP, fill=X)
        self.edit_title.pack(fill=X)
        self.edit_entry_frame.pack(fill=X)
        self.edit_btn_frame.pack(side=BOTTOM, fill=X)

        # Packing all entries >>
        self.edit_name_frame.pack(fill=X)
        self.edit_name_label.pack(side=LEFT)
        self.edit_name.pack(side=RIGHT)

        self.edit_date_frame.pack(fill=X)
        self.edit_date_label.pack(side=LEFT)
        self.edit_date.pack(side=RIGHT)

        self.edit_amount_frame.pack(fill=X)
        self.edit_amount_label.pack(side=LEFT)
        self.edit_amount.pack(side=RIGHT)

        self.edit_place_frame.pack(fill=X)
        self.edit_place_label.pack(side=LEFT)
        self.edit_place.pack(side=RIGHT)

        self.edit_phone_frame.pack(fill=X)
        self.edit_phone_label.pack(side=LEFT)
        self.edit_phone.pack(side=RIGHT)

        # Packing buttons for editing data >>
        self.edit_btn_frame1.pack(side=TOP, fill=X)
        self.edit_btn_frame2.pack(fill=X)
        self.edit_btn_frame3.pack(side=BOTTOM, fill=X)
        self.edit_btn_edit.pack(side=LEFT)
        self.edit_btn_history.place(anchor=CENTER, x=220, y=19)
        self.edit_btn_delete.pack(side=RIGHT)
        self.edit_credit_frame.pack(side=TOP, fill=X)
        self.edit_debit_frame.pack(side=BOTTOM, fill=X)
        self.edit_btn_credit.pack(fill=X)
        self.edit_btn_debit.pack(side=BOTTOM, fill=X)
        self.edit_btn_new_entry.pack(fill=X)

        # Disable all entries >>
        self.edit_name['state'] = DISABLED
        self.edit_date['state'] = DISABLED
        self.edit_amount['state'] = DISABLED
        self.edit_place['state'] = DISABLED
        self.edit_phone['state'] = DISABLED
        self.edit_btn_edit['state'] = DISABLED
        self.edit_btn_delete['state'] = DISABLED
        self.edit_btn_credit['state'] = DISABLED
        self.edit_btn_debit['state'] = DISABLED
        self.edit_btn_history['state'] = DISABLED


class Run:

    @staticmethod
    def run():
        app = Main()
        app.make_title()
        app.make_search_window()
        app.make_edit_frame()
        app.edit_tab()
        app.mainloop()


if __name__ == '__main__':
    try:
        Check().check_all()
        Run.run()
    except Exception as e11:
        Main.show_error(e11)
